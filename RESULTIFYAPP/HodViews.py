from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render
from django.urls import reverse
from django.shortcuts import get_object_or_404
from openpyxl.reader.excel import load_workbook

from RESULTIFYAPP.forms import AddStudentForm, EditStudentForm, FileUploadForm
from RESULTIFYAPP.models import CustomUser, Staffs, Department, Courses, Students, Assessment


def admin_home(request):
    return render(request,"hod_template/home_content.html")

def add_staff(request):
    return render(request,"hod_template/add_staff_template.html")


def add_staff_save(request):
    if request.method!="POST":
        return HttpResponse("Method Not Allowed")
    else:
        first_name=request.POST.get("first_name")
        last_name=request.POST.get("last_name")
        username=request.POST.get("username")
        email=request.POST.get("email")
        password=request.POST.get("password")
        address=request.POST.get("address")
        try:
            user=CustomUser.objects.create_user(username=username,password=password,email=email,last_name=last_name,first_name=first_name,user_type=2)
            user.staffs.address=address
            user.save()
            messages.success(request,"Successfully Added Staff")
            return HttpResponseRedirect(reverse("add_staff"))
        except:
            messages.error(request,"Failed to Add Staff")
            return HttpResponseRedirect(reverse("add_staff"))

def add_department(request):
    return render(request, "hod_template/add_department_template.html")

def add_department_save(request):
    if request.method!="POST":
        return HttpResponse("Method Not Allowed")
    else:
        department_name = request.POST.get("department")
        try:
            department_model = Department(department_name=department_name)
            department_model.save()
            messages.success(request, "Successfully Added Department")
            return HttpResponseRedirect(reverse("add_department"))
        except Exception as e:
            print(f"Error while adding department: {str(e)}")
            messages.error(request, "Failed To Add Department")
            return HttpResponseRedirect(reverse("add_department"))


def add_course(request):
    return render(request,"hod_template/add_course_template.html")

def add_course_save(request):
    if request.method != "POST":
        return HttpResponseRedirect("Method Not Allowed")
    else:
        course = request.POST.get("course")
        try:
            course_model = Courses(course_name=course)
            course_model.save()
            messages.success(request, "Successfully Added Course")
            return HttpResponseRedirect(reverse("add_course"))
        except Exception as e:
            print(f"Error while adding course: {str(e)}")
            messages.error(request, "Failed To Add Course")
            return HttpResponseRedirect(reverse("add_course"))
def add_student(request):
    form=AddStudentForm()
    return render(request,"hod_template/add_student_template.html",{"form":form})

def add_student_save(request):
    if request.method!="POST":
        return HttpResponse("Method Not Allowed")
    else:
        form=AddStudentForm(request.POST,request.FILES)
        if form.is_valid():
            first_name=form.cleaned_data["first_name"]
            last_name=form.cleaned_data["last_name"]
            username=form.cleaned_data["username"]
            matric_number=form.cleaned_data["matric_number"]
            email=form.cleaned_data["email"]
            password=form.cleaned_data["password"]
            address=form.cleaned_data["address"]
            session_start=form.cleaned_data["session_start"]
            session_end=form.cleaned_data["session_end"]
            # course_id=form.cleaned_data["course"]
            sex=form.cleaned_data["sex"]

            profile_pic=request.FILES['profile_pic']
            fs=FileSystemStorage()
            filename=fs.save(profile_pic.name,profile_pic)
            profile_pic_url=fs.url(filename)

            try:
                user=CustomUser.objects.create_user(username=username,password=password,email=email,last_name=last_name,first_name=first_name,user_type=3)
                user.students.address=address
                user.students.matric_number=matric_number
                # course_obj=Courses.objects.get(course_code='ECE 6010')
                # user.students.course_id=course_obj
                user.students.session_start_year=session_start
                user.students.session_end_year=session_end
                user.students.gender=sex
                user.students.profile_pic=profile_pic_url
                user.save()
                messages.success(request,"Successfully Added Student")
                return HttpResponseRedirect(reverse("add_student"))
            except:
                messages.error(request,"Failed to Add Student")
                return HttpResponseRedirect(reverse("add_student"))
        else:
            form=AddStudentForm(request.POST)
            return render(request, "hod_template/add_student_template.html", {"form": form})

# def add_assessment(request):
#     return render(request, "hod_template/add_assessment_template.html")


def add_assessment(request):

    if request.method == 'POST':
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file_uploaded = form.cleaned_data['file']
            session = file_uploaded.name.split('.')[0]
            workbook = load_workbook(file_uploaded)
            for sheet_name in workbook.sheetnames:
                sheet_name = str(sheet_name).replace('_', '/')
                for index, row in enumerate(workbook.active.iter_rows(values_only=True)):
                    if index == 0:
                        continue
                    score = int(row[2])
                    if score >= 70:
                        grade = 'A'
                    elif score >= 60:
                        grade = 'B'
                    elif score >= 50:
                        grade = 'C'
                    elif score >= 40:
                        grade = 'D'
                    elif score >= 30:
                        grade = 'E'
                    else:
                        grade = 'F'
                    try:
                        course = Courses.objects.get(course_code=sheet_name)
                    except Courses.DoesNotExist:
                        department = Department.objects.get(id=1)
                        course = department.courses_set.create(
                            course_name="Dummy Course Name",
                            credit_unit=0, course_code=sheet_name,
                        )
                    course.assessment_set.create(full_name=row[0], mat_number=row[1],
                        score=score, grade=grade, session=session)
            return HttpResponseRedirect(reverse("success_assessment"))
        return HttpResponse('An error occurred')

    return render(request, 'hod_template/add_assessment_template.html', {'form': FileUploadForm()})
 #   else:
  #      form = FileUploadForm()
   # return render(request, 'hod_template/add_assessment_template.html', {'form': form})
def manage_staff(request):
    staffs=Staffs.objects.all()
    return render(request,"hod_template/manage_staff_template.html",{"staffs":staffs})
def manage_department(request):
    departments=Department.objects.all()
    return render(request, "hod_template/manage_department_template.html", {"departments":departments})

def manage_student(request):
    students=Students.objects.all()
    return render(request,"hod_template/manage_student_template.html",{"students":students})

def manage_course(request):
    courses=Courses.objects.all()
    return render(request,"hod_template/manage_course_template.html",{"courses":courses})

def manage_assessment(request):
    assessments=Assessment.objects.all()
    return render(request,"hod_template/manage_assessment_template.html",{"assessments":assessments})

def edit_staff(request,staff_id):
    staff=Staffs.objects.get(admin=staff_id)
    return render(request,"hod_template/edit_staff_template.html",{"staff":staff,"id":staff_id})

def edit_staff_save(request):
    if request.method!="POST":
        return HttpResponse("<h2>Method Not Allowed</h2>")
    else:
        staff_id=request.POST.get("staff_id")
        first_name=request.POST.get("first_name")
        last_name=request.POST.get("last_name")
        email=request.POST.get("email")
        username=request.POST.get("username")
        address=request.POST.get("address")

        try:
            user=CustomUser.objects.get(id=staff_id)
            user.first_name=first_name
            user.last_name=last_name
            user.email=email
            user.username=username
            user.save()

            staff_model=Staffs.objects.get(admin=staff_id)
            staff_model.address=address
            staff_model.save()
            messages.success(request,"Successfully Edited Staff")
            return HttpResponseRedirect(reverse("edit_staff",kwargs={"staff_id":staff_id}))
        except:
            messages.error(request,"Failed to Edit Staff")
            return HttpResponseRedirect(reverse("edit_staff",kwargs={"staff_id":staff_id}))

def edit_department(request, department_id):
    department = Department.objects.get(id=department_id)
    return render(request, "hod/edit_department_template.html", {"department": department, "id": department_id})

def edit_department_save(request):
    if request.method!="POST":
        return HttpResponse("<h2>Method Not Allowed</h2>")
    else:
        department_id = request.POST.get("department_id")
        department_name = request.POST.get("department")
        try:
            department = Department.objects.get(id=department_id)
            department.department_name = department_name
            department.save()
            messages.success(request, "Successfully Edited Department")
            return HttpResponseRedirect(reverse("edit_department", kwargs={"department_id": department_id}))
        except:
            messages.error(request, "Failed to Edit Department")
            return HttpResponseRedirect(reverse("edit_department", kwargs={"department_id": department_id}))


def edit_student(request,student_id):
    request.session['student_id']=student_id
    student=Students.objects.get(admin=student_id)
    form=EditStudentForm()
    form.fields['email'].initial=student.admin.email
    form.fields['first_name'].initial=student.admin.first_name
    form.fields['last_name'].initial=student.admin.last_name
    form.fields['username'].initial=student.admin.username
    form.fields['address'].initial=student.address
    form.fields['course'].initial=student.course_id.id
    form.fields['sex'].initial=student.gender
    form.fields['session_start'].initial=student.session_start_year
    form.fields['session_end'].initial=student.session_end_year
    return render(request,"hod_template/edit_student_template.html",{"form":form,"id":student_id,"username":student.admin.username})

def edit_student_save(request):
    if request.method!="POST":
        return HttpResponse("<h2>Method Not Allowed</h2>")
    else:
        student_id=request.session.get("student_id")
        if student_id==None:
            return HttpResponseRedirect(reverse("manage_student"))

        form=EditStudentForm(request.POST,request.FILES)
        if form.is_valid():
            first_name = form.cleaned_data["first_name"]
            last_name = form.cleaned_data["last_name"]
            username = form.cleaned_data["username"]
            email = form.cleaned_data["email"]
            address = form.cleaned_data["address"]
            session_start = form.cleaned_data["session_start"]
            session_end = form.cleaned_data["session_end"]
            course_id = form.cleaned_data["course"]
            sex = form.cleaned_data["sex"]

            if request.FILES.get('profile_pic',False):
                profile_pic=request.FILES['profile_pic']
                fs=FileSystemStorage()
                filename=fs.save(profile_pic.name,profile_pic)
                profile_pic_url=fs.url(filename)
            else:
                profile_pic_url=None


            try:
                user=CustomUser.objects.get(id=student_id)
                user.first_name=first_name
                user.last_name=last_name
                user.username=username
                user.email=email
                user.save()

                student=Students.objects.get(admin=student_id)
                student.address=address
                student.session_start_year=session_start
                student.session_end_year=session_end
                student.gender=sex
                course=Courses.objects.get(id=course_id)
                student.course_id=course
                if profile_pic_url!=None:
                    student.profile_pic=profile_pic_url
                student.save()
                del request.session['student_id']
                messages.success(request,"Successfully Edited Student")
                return HttpResponseRedirect(reverse("edit_student",kwargs={"student_id":student_id}))
            except:
                messages.error(request,"Failed to Edit Student")
                return HttpResponseRedirect(reverse("edit_student",kwargs={"student_id":student_id}))
        else:
            form=EditStudentForm(request.POST)
            student=Students.objects.get(admin=student_id)
            return render(request,"hod_template/edit_student_template.html",{"form":form,"id":student_id,"username":student.admin.username})


def edit_course(request, course_id):
    # Get the existing course using get_object_or_404
    course = get_object_or_404(Courses, id=course_id)

    # Now you can retrieve the department and staff associated with the course
    department = course.department_id
    staff = course.staff_id

    return render(request, "hod_template/edit_course_template.html", {
        "course": course,
        "id": course_id,
        "department": department,
        "staff": staff
    })
#def edit_course(request,course_id):
#    course=Courses.objects.get(id=course_id)
#    return render(request,"hod_template/edit_course_template.html",{"course":course,"id":course_id})

def edit_course_save(request):
    if request.method != "POST":
        return HttpResponse("<h2>Method Not Allowed</h2>")
    else:
        course_id = request.POST.get("course_id")
        course_name = request.POST.get("course")
        department_id = request.POST.get("department")  # Retrieve the selected department
        staff_id = request.POST.get("staff")  # Retrieve the selected staff
        credit_unit = request.POST.get("credit_unit")
        course_code = request.POST.get("course_code")

        try:
            course = Courses.objects.get(id=course_id)
            course.course_name = course_name
            course.department_id_id = department_id  # Set the department using its ID
            course.staff_id_id = staff_id  # Set the staff using its ID
            course.credit_unit = credit_unit
            course.course_code = course_code
            course.save()
            messages.success(request, "Successfully Edited Course")
            return HttpResponseRedirect(reverse("edit_course", kwargs={"course_id": course_id}))
        except:
            messages.error(request, "Failed to Edit Course")
            return HttpResponseRedirect(reverse("edit_course", kwargs={"course_id": course_id}))
def edit_assessment(request, assessment_id):
    assessment = Assessment.objects.get(id=assessment_id)
    return render(request, "hod_template/edit_assessment_template.html", {"assessment": assessment, "id": assessment_id})


def edit_assessment_save(request):
    if request.method != "POST":
        return HttpResponse("<h2>Method Not Allowed</h2>")
    else:
        assessment_id = request.POST.get("assessment_id")
        name = request.POST.get("name")
        matric_number = request.POST.get("matric_number")
        robotics = request.POST.get("robotics")
        software_engineering = request.POST.get("software_engineering")
        verilog = request.POST.get("verilog")
        total_score = request.POST.get("total_score")
        grades = request.POST.get("grades")
        bgs = request.POST.get("bgs")

        # Fetch the existing assessment using get_object_or_404
        assessment = get_object_or_404(Assessment, id=assessment_id)

        try:
            # Update the assessment fields
            assessment.name = name
            assessment.matric_number = matric_number
            assessment.robotics = robotics
            assessment.software_engineering = software_engineering
            assessment.verilog = verilog
            assessment.total_score = total_score
            assessment.grades = grades
            assessment.bgs = bgs
            assessment.save()  # Save the updated assessment

            messages.success(request, "Successfully Updated Assessment")
            return HttpResponseRedirect(reverse("add_assessment"))
        except Exception as e:
            print(f"Error while updating assessment: {str(e)}")
            messages.error(request, "Failed To Update Assessment")
            return HttpResponseRedirect(reverse("add_assessment"))

def success_assessment(request):
    return render(request, "hod_template/success_add_assessment.html")